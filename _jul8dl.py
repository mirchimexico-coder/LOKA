# -*- coding: utf-8 -*-
import openpyxl
from copy import copy
from datetime import date
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
wb=openpyxl.load_workbook(P); dl=wb.worksheets[2]
MONEY='$#,##0.00'
# ---- Soft card columns + rate cell ----
dl.cell(7,21,'Soft Rest Card $').font=copy(dl.cell(7,18).font)
dl.cell(7,22,'Soft Comm $').font=copy(dl.cell(7,18).font)
dl.cell(7,23,'Soft Rate').font=copy(dl.cell(7,19).font)
xc=dl.cell(7,24,0.0205); xc.number_format='0.00%'; xc.font=copy(dl.cell(7,20).font)
# ---- July 8 row 58 ----
r=58; tr=57  # template row (Jul 7)
# copy formatting from template row across cols 2..24
for c in range(2,25):
    d=dl.cell(r,c); srcc=dl.cell(tr,c)
    d.font=copy(srcc.font); d.fill=copy(srcc.fill); d.border=copy(srcc.border); d.alignment=copy(srcc.alignment); d.number_format=srcc.number_format
d=date(2026,7,8)
dl.cell(r,2,d); dl.cell(r,2).number_format='dd-mmm-yyyy'
dl.cell(r,3,d.strftime('%a'))
dl.cell(r,4,360)     # MP card
dl.cell(r,5,2055)    # cash
dl.cell(r,7,195)     # transfer-to-me
dl.cell(r,21,1078)   # Soft Restaurant card
dl.cell(r,6, f'=IFERROR(D{r}+E{r}+G{r}+U{r},0)')            # total revenue incl soft
# copy H (expenses SUMIFS) from template, adjust row refs
hf=dl.cell(tr,8).value
dl.cell(r,8, hf.replace(str(tr),str(r)) if isinstance(hf,str) and hf.startswith('=') else f'=SUMIFS(\'\U0001f4b8 Expenses\'!$F$8:$F$500,\'\U0001f4b8 Expenses\'!$B$8:$B$500,">="&B{r},\'\U0001f4b8 Expenses\'!$B$8:$B$500,"<"&(B{r}+1))')
dl.cell(r,9, f'=IFERROR(F{r}-H{r},0)')                      # net
dl.cell(r,10, f'=IFERROR(I{r}/F{r},0)')                     # margin
dl.cell(r,18, f'=IFERROR(D{r}*$T$7,0)')                     # MP comm
dl.cell(r,22, f'=IFERROR(U{r}*$X$7,0)')                     # Soft comm
dl.cell(r,18).number_format=MONEY; dl.cell(r,22).number_format=MONEY; dl.cell(r,21).number_format=MONEY
dl.cell(r,12,'EOD close (dual card: MP + Soft Restaurant)')
wb.calculation.fullCalcOnLoad=True
wb.save(P)
print("Jul8 row 58 added. day=",d.strftime('%a'))
print("H formula:", dl.cell(r,8).value)
print("MP comm expect", round(360*0.0406,2), "| Soft comm expect", round(1078*0.0205,2))
print("revenue expect", 360+2055+195+1078)
