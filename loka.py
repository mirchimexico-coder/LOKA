#!/usr/bin/env python
# -*- coding: utf-8 -*-
# C:\LOKA\loka.py — LOKA bookkeeping toolkit (reusable). Never use edit_block on the xlsx.
import sys, os, json, argparse
from datetime import datetime, date
import datetime as dtmod
from copy import copy
import openpyxl

P = r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
sys.path.insert(0, r'C:\LOKA')
S_DAILY, S_EXP = 2, 3
OWNER = '💰 Owner Ledger'
MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def backup(label='auto'):
    from backup_util import backup_excel
    return backup_excel(label)

def pdate(s):
    if isinstance(s, (datetime, date)):
        return s.date() if isinstance(s, datetime) else s
    s = str(s).strip()
    for fmt in ('%d-%b-%Y', '%d-%b', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            d = dtmod.datetime.strptime(s, fmt).date()
            if fmt == '%d-%b':
                d = d.replace(year=date.today().year)
            return d
        except ValueError:
            continue
    raise ValueError('Bad date: ' + s)

def _last_data_row(ws, key_col=3, start=8):
    r = ws.max_row
    while r > start and ws.cell(r, key_col).value is None:
        r -= 1
    return r

def _copyfmt(ws, src_row, dst_row, cols):
    for c in cols:
        s, d = ws.cell(src_row, c), ws.cell(dst_row, c)
        d.number_format = s.number_format
        d.font = copy(s.font); d.border = copy(s.border); d.alignment = copy(s.alignment)

def add_expenses(rows, do_backup=True):
    """rows: list of dict(date,desc,vendor,cat,amount,paid,method,notes)."""
    if do_backup: backup('add_expenses')
    wb = openpyxl.load_workbook(P); ex = wb.worksheets[S_EXP]
    last = _last_data_row(ex)
    n = 0
    for i, r in enumerate(rows):
        tr = last + 1 + i
        vals = [pdate(r['date']), r['desc'], r.get('vendor','Local purchase'),
                r['cat'], float(r['amount']), r.get('paid','Restaurant'),
                r.get('method','Cash'), r.get('status','✅ Paid'), r.get('notes','')]
        for c, v in enumerate(vals, start=2):
            ex.cell(tr, c, v)
        ex.cell(tr, 2).number_format = 'dd-mmm-yyyy'
        _copyfmt(ex, last, tr, range(2,11))
        ex.cell(tr,2).number_format = 'dd-mmm-yyyy'
        n += 1
    wb.save(P)
    return n, last+1, last+n

def _inject_cache():
    """Write computed values into the <v> cache of Daily Log formula cells.

    openpyxl writes formulas with an EMPTY <v/> cache, and Excel here does not
    reliably honour fullCalcOnLoad -> cells render blank until a manual recalc.
    This rewrites xl/worksheets/sheet3.xml keeping <f> intact and filling <v>.
    MUST run after any write that touches the Daily Log (refresh_dashboard does
    it automatically at the end).
    """
    import zipfile, shutil, re as _re
    from datetime import timedelta
    wb = openpyxl.load_workbook(P); dl = wb.worksheets[S_DAILY]; ex = wb.worksheets[S_EXP]
    def _pd(v): return v.date() if isinstance(v, datetime) else (v if isinstance(v, date) else None)
    lastx = _last_data_row(ex)
    ebd = {}
    for r in range(8, lastx+1):
        dd = _pd(ex.cell(r,2).value); a = ex.cell(r,6).value
        if isinstance(dd, date) and isinstance(a,(int,float)): ebd[dd] = ebd.get(dd,0.0)+float(a)
    vals = {}; today_row = None
    _dates=[]; _revbd={}; _tfbd={}
    for r in range(8, dl.max_row+1):
        dd = _pd(dl.cell(r,2).value)
        if not isinstance(dd, date): continue
        rev = sum(float(dl.cell(r,c).value or 0) for c in (4,5,7,21,25))
        e = round(ebd.get(dd,0.0),2); net = round(rev-e,2)
        _dates.append(dd); _revbd[dd]=_revbd.get(dd,0.0)+rev
        _tfbd[dd]=_tfbd.get(dd,0.0)+float(dl.cell(r,7).value or 0)
        vals[f'F{r}']=rev; vals[f'H{r}']=e; vals[f'I{r}']=net
        vals[f'J{r}']=round(net/rev,6) if rev else 0
        vals[f'R{r}']=round(float(dl.cell(r,4).value or 0)*CFG['commission_rate'],2)
        today_row = (rev, e, net)
    if today_row and _dates:           # top-of-sheet "today / this week / this month" row 6
        rev,e,net = today_row
        vals['H6']=e; vals['I6']=net; vals['J6']=round(net/rev,6) if rev else 0
        # F6/G6 = today's figures; M6/N6/O6/P6/Q6 = week & month roll-ups.
        # "Today" here means the LATEST recorded day (the sheet uses TODAY(); if no
        # trading happened today these legitimately show the last day's numbers).
        td = max(d for d in _dates)
        vals['F6'] = round(_revbd.get(td,0.0),2)
        vals['G6'] = round(_tfbd.get(td,0.0),2)
        wk_start = td - timedelta(days=(td.weekday()-0) % 7)   # weeks start Monday
        wk = [d for d in _dates if wk_start <= d < wk_start+timedelta(days=7)]
        mo = [d for d in _dates if d.month==td.month and d.year==td.year]
        wk_rev = round(sum(_revbd.get(d,0.0) for d in wk),2)
        wk_exp = round(sum(ebd.get(d,0.0) for d in wk),2)
        vals['M6']=wk_rev; vals['N6']=wk_rev
        vals['O6']=round(sum(_revbd.get(d,0.0) for d in mo),2)
        vals['P6']=wk_exp; vals['Q6']=round(wk_rev-wk_exp,2)
    tmp = P+'.tmp'
    zin = zipfile.ZipFile(P,'r'); zout = zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED)
    n = 0
    for item in zin.namelist():
        data = zin.read(item)
        if item == 'xl/worksheets/sheet3.xml':
            xml = data.decode('utf-8')
            def _repl(m):
                nonlocal n
                ref = m.group('ref'); whole = m.group(0)
                if ref in vals and '<f' in whole:
                    nv = f"<v>{vals[ref]}</v>"
                    w = _re.sub(r'<v\s*/>', nv, whole)
                    if w == whole: w = _re.sub(r'<v>.*?</v>', nv, whole)
                    if '<v' not in w: w = w.replace('</c>', nv+'</c>')
                    if w != whole: n += 1
                    return w
                return whole
            # NOTE: the alternation is essential - a self-closing cell like
            # <c r="Q74" s="1"/> would otherwise let .*?</c> run on and swallow the
            # NEXT cell, so that cell never gets its cached value (this silently
            # left the whole R/MP-commission column blank until 27-Jul).
            xml = _re.sub(r'<c r="(?P<ref>[A-Z]+\d+)"[^>]*?(?:/>|>.*?</c>)', _repl, xml)
            data = xml.encode('utf-8')
        zout.writestr(item, data)
    zin.close(); zout.close(); shutil.move(tmp, P)
    return n

def close_day(d, card=0, cash=0, transfer=0, soft=0, softcomm=0, bbva=0, bbvacomm=0, do_backup=True):
    """Set revenue for date d in Daily Log; create row with auto-SUMIFS if missing."""
    d = pdate(d)
    if do_backup: backup('close_day')
    wb = openpyxl.load_workbook(P); dl = wb.worksheets[S_DAILY]
    EXP = "'💸 Expenses'"
    # find row with this date
    target = None; lastrow = 7
    for r in range(8, dl.max_row+1):
        v = dl.cell(r,2).value
        if v is None: continue
        lastrow = r
        dd = v.date() if isinstance(v, datetime) else v
        if isinstance(dd, date) and dd == d:
            target = r; break
    if target is None:
        target = lastrow + 1
        dl.cell(target,2, d); dl.cell(target,2).number_format='dd-mmm-yyyy'
        dl.cell(target,3, MONTHS[d.month-1] and d.strftime('%a'))
        dl.cell(target,6, f'=IFERROR(D{target}+E{target}+G{target}+U{target}+Y{target},0)')
        dl.cell(target,8, f"=SUMIFS({EXP}!$F$8:$F$5000,{EXP}!$B$8:$B$5000,\">=\"&B{target},{EXP}!$B$8:$B$5000,\"<\"&(B{target}+1))")
        dl.cell(target,9, f'=IFERROR(F{target}-H{target},0)')
        dl.cell(target,10, f'=IFERROR(I{target}/F{target},0)')
        dl.cell(target,18, f'=IFERROR(D{target}*$T$7,0)'); dl.cell(target,18).number_format='$#,##0.00'
        _copyfmt(dl, lastrow, target, range(2,29))
        dl.cell(target,2).number_format='dd-mmm-yyyy'
        for cc in (18,21,22,25,26): dl.cell(target,cc).number_format='$#,##0.00'
    dl.cell(target,4, float(card)); dl.cell(target,5, float(cash)); dl.cell(target,7, float(transfer))
    dl.cell(target,21, float(soft)); dl.cell(target,22, float(softcomm))
    dl.cell(target,25, float(bbva)); dl.cell(target,26, float(bbvacomm))
    dl.cell(target,12,'Cierre via loka.py')
    wb.save(P)
    return target, float(card)+float(cash)+float(transfer)+float(soft)+float(bbva)

def add_ledger(d, desc, spent=0, transferred=0, typ=None, status=None, notes='', do_backup=True):
    d = pdate(d)
    if do_backup: backup('owner_ledger')
    wb = openpyxl.load_workbook(P); ol = wb[OWNER]
    # find TOTALS row
    tr = None
    for r in range(4, ol.max_row+1):
        if str(ol.cell(r,1).value or '').startswith('TOTALS'):
            tr = r; break
    last = tr - 1            # last data row
    new = tr                 # entry goes where TOTALS was
    _copyfmt(ol, last, new, range(1,9))
    for c in range(1,9): ol.cell(new,c).value=None
    ol.cell(new,1, d); ol.cell(new,1).number_format='dd-mmm-yyyy'
    ol.cell(new,2, desc); ol.cell(new,3, typ or ('Transfer Received' if transferred else 'Expense - Personal'))
    if spent: ol.cell(new,4, float(spent))
    if transferred: ol.cell(new,5, float(transferred))
    ol.cell(new,6, f'=IFERROR(F{last}+D{new}-E{new},0)')
    ol.cell(new,7, status or ('✅ Received' if transferred else '⏳ Reimburse'))
    ol.cell(new,8, notes)
    # rewrite TOTALS one row down
    t = new + 1
    _copyfmt(ol, tr, t, range(1,9))
    ol.cell(t,1,'TOTALS & CURRENT BALANCE')
    ol.cell(t,4, f'=SUM(D4:D{new})'); ol.cell(t,5, f'=SUM(E4:E{new})'); ol.cell(t,6, f'=F{new}')
    ol.cell(t,7, f'=IF(F{new}>0,"Restaurant owes Lohith $"&TEXT(F{new},"#,##0"),IF(F{new}<0,"Lohith holds $"&TEXT(ABS(F{new}),"#,##0"),"Zero balance"))')
    wb.save(P)
    return new

def cash_adjust_add(delta, reason):
    """Append a delta to CFG['cash_adjust'] in this file and log the reason.

    Use for anything that moves till cash but is NOT a Daily Log expense:
      transfer-to-me      -> negative (cash went to Lohith, not the till)
      owner-paid expense  -> positive (till never paid it; add back)
      capital repayment   -> negative (cash left the restaurant)
    """
    import re as _re
    src = open(__file__, encoding='utf-8').read()
    m = _re.search(r'(cash_adjust=)(-?[\d.]+)(,\s*#[^\n]*)', src)
    if not m: raise SystemExit('cash_adjust line not found')
    old = float(m.group(2)); new = round(old + float(delta), 2)
    note = m.group(3).rstrip()
    stamp = f" | {date.today():%d-%b}: {'+' if float(delta)>=0 else ''}{float(delta):,.2f} {reason}"
    src = src[:m.start()] + f"cash_adjust={new:.2f}" + note + stamp + src[m.end():]
    open(__file__, 'w', encoding='utf-8', newline='\n').write(src)
    # IMPORTANT: also update the value loaded in THIS process, otherwise a
    # refresh_all() called afterwards in the same run recomputes cash with the
    # stale figure (this overstated Cash on Hand by the transfer amount on 27-Jul).
    CFG['cash_adjust'] = new
    return old, new

def refresh_all(do_backup=True):
    """One-shot: P&L sync + dashboard + formula-cache injection."""
    refresh_dashboard(do_backup=do_backup)
    n = _inject_cache()
    print(f'  cache injected into {n} cells')

def compute():
    """Summary for the `status` command.

    Delegates to _gather() ON PURPOSE. This used to be a second, parallel
    implementation that read only MP/cash/transfer -- it silently ignored the Soft
    and BBVA cards and understated all-time revenue by $25,780, and it still used a
    hard-coded +20000 for the rent add-back. Never re-implement the maths here.
    """
    g = _gather()
    days = g['days']
    soft_t = round(sum(g['soft_bd'].values()), 2)
    bbva_t = round(sum(g['bbva_bd'].values()), 2)
    rev = round(sum(c+k+t for _,c,k,t,_ in days) + soft_t + bbva_t, 2)
    exp = round(sum(g['ebd'].values()), 2)
    comm = round(sum(c for _,c,_,_,_ in days)*CFG['commission_rate']
                 + sum(g['softcomm_bd'].values()) + sum(g['bbvacomm_bd'].values()), 2)
    trading = sum(1 for _,c,k,t,_ in days if c+k+t > 0)
    # cash on hand, exactly as refresh_dashboard computes it
    A = CFG['cash_anchor_date']
    roll = sum((c+k+t)-(e-g['capf_bd'].get(d,0.0)) for d,c,k,t,e in days if d > A)
    soft_p = round(sum(v for dd,v in g['soft_bd'].items() if dd > A), 2)
    bbva_p = round(sum(v for dd,v in g['bbva_bd'].items() if dd > A), 2)
    comm_p = (sum(c*CFG['commission_rate'] for d,c,_,_,_ in days if d > A)
              + sum(v for dd,v in g['softcomm_bd'].items() if dd > A)
              + sum(v for dd,v in g['bbvacomm_bd'].items() if dd > A))
    cash = round(CFG['cash_anchor_amount'] + roll + soft_p + bbva_p - comm_p + CFG['cash_adjust'], 2)
    out = {'all_time': {
             'revenue': rev, 'expenses': exp, 'commission': comm,
             'net': round(rev-exp, 2),
             'net_after_commission': round(rev-exp-comm, 2),
             'operating_net': round(rev-exp+g['owe_capital'], 2),
             'trading_days': trading,
             'card_mp': round(sum(d[1] for d in days),2), 'card_soft': soft_t, 'card_bbva': bbva_t,
             'cash_taken': round(sum(d[2] for d in days),2),
             'transfer': round(sum(d[3] for d in days),2)},
           'position': {
             'cash_on_hand': cash,
             'owed_to_capital': g['owe_capital'],
             'owner_ledger': g['ol_balance'],
             'net_cash_position': round(cash - g['owe_capital'] - g['ol_balance'], 2),
             'cash_anchor': f"{CFG['cash_anchor_date']} = {CFG['cash_anchor_amount']:,.2f}"},
           'by_month': {},
           'last_days': [(str(d), round(c+k+t+g['soft_bd'].get(d,0)+g['bbva_bd'].get(d,0),2), e,
                          round(c+k+t+g['soft_bd'].get(d,0)+g['bbva_bd'].get(d,0)-e,2))
                         for d,c,k,t,e in days[-7:]]}
    for k in sorted(g['inc_m']):
        ti = sum(g['inc_m'][k].values()); te = sum(g['cat_m'][k].values())
        out['by_month'][f'{MONTHS[k[1]-1]} {k[0]}'] = {
            'income': round(ti,2), 'expenses': round(te,2), 'net': round(ti-te,2),
            'days': g['dcount'][k],
            'top_cats': sorted(((c,round(a,2)) for c,a in g['cat_m'][k].items()), key=lambda x:-x[1])[:6]}
    return out

def main():
    ap = argparse.ArgumentParser(description='LOKA bookkeeping toolkit')
    sub = ap.add_subparsers(dest='cmd')
    sub.add_parser('status')
    sub.add_parser('backup').add_argument('label', nargs='?', default='manual')
    a = sub.add_parser('add-expense')
    for x in ['date','desc','vendor','cat','amount']: a.add_argument('--'+x, required=True)
    a.add_argument('--paid', default='Restaurant'); a.add_argument('--method', default='Cash'); a.add_argument('--notes', default='')
    c = sub.add_parser('close-day')
    c.add_argument('--date', required=True); c.add_argument('--card', default=0); c.add_argument('--cash', default=0); c.add_argument('--transfer', default=0)
    c.add_argument('--soft', default=0); c.add_argument('--softcomm', default=0)
    c.add_argument('--bbva', default=0); c.add_argument('--bbvacomm', default=0)
    c.add_argument('--bbvarate', default=0.019, type=float, help='if --bbvacomm omitted, comm = bbva*rate')
    l = sub.add_parser('add-ledger')
    l.add_argument('--date', required=True); l.add_argument('--desc', required=True)
    l.add_argument('--spent', default=0); l.add_argument('--transferred', default=0)
    l.add_argument('--typ', default=None); l.add_argument('--status', default=None); l.add_argument('--notes', default='')
    ca = sub.add_parser('cash-adjust')
    ca.add_argument('--delta', required=True); ca.add_argument('--reason', required=True)
    b = sub.add_parser('add-expenses-json'); b.add_argument('file')
    sub.add_parser('refresh-dashboard')
    sub.add_parser('refresh-all')
    sub.add_parser('inject-cache')
    sub.add_parser('refresh-pl')
    args = ap.parse_args()
    if args.cmd=='status': print(json.dumps(compute(), indent=2, ensure_ascii=False))
    elif args.cmd=='refresh-dashboard': refresh_dashboard()
    elif args.cmd=='refresh-all': refresh_all()
    elif args.cmd=='inject-cache': print(f'cache injected into {_inject_cache()} cells')
    elif args.cmd=='refresh-pl': print(refresh_pl())
    elif args.cmd=='backup': print(backup(args.label))
    elif args.cmd=='add-expense':
        n,a0,a1=add_expenses([{'date':args.date,'desc':args.desc,'vendor':args.vendor,'cat':args.cat,'amount':args.amount,'paid':args.paid,'method':args.method,'notes':args.notes}])
        print(f'added rows {a0}-{a1}')
    elif args.cmd=='add-expenses-json':
        n,a0,a1=add_expenses(json.load(open(args.file,encoding='utf-8'))); print(f'added {n} rows {a0}-{a1}')
    elif args.cmd=='add-ledger':
        r=add_ledger(args.date,args.desc,args.spent,args.transferred,args.typ,args.status,args.notes)
        print(f'ledger row {r}')
    elif args.cmd=='cash-adjust':
        o,n=cash_adjust_add(args.delta,args.reason); print(f'cash_adjust {o} -> {n}')
    elif args.cmd=='close-day':
        bc = float(args.bbvacomm) or round(float(args.bbva)*float(args.bbvarate),2)
        row,tot=close_day(args.date,args.card,args.cash,args.transfer,args.soft,args.softcomm,args.bbva,bc)
        print(f'row {row} revenue {tot} (bbva comm {bc})')
    else: ap.print_help()
# (entry point moved to end of file, after all helper defs)


# ============== DASHBOARD REFRESH ==============
DASH = r'C:\LOKA\dashboard.html'
WEEK_COLORS = ['#3b82f6','#22c55e','#f97316','#a855f7','#f59e0b','#ef4444','#06b6d4','#ec4899']
# Manual anchors that change rarely — update here when the situation changes.
CFG = dict(
    cash_anchor_date=date(2026,7,26), cash_anchor_amount=15395.0,
    cash_adjust=-848.00,                   # RESET 26-Jul-2026: full physical count of all cash forms (till + bank/card balances) = $15,395 became the new anchor. The fresh count absorbs ALL prior drift, the Jun29 withdrawal, transfers-to-Lohith, owner-paid add-backs, and the $22,587 Capital repayments. Start clean from here: only add NEW adjustments dated AFTER 26-Jul (transfers-to-Lohith, owner-paid expenses, capital repayments). MP commission auto-deducted per day. | 27-Jul: +0.00 automation smoke test | 27-Jul: -700.00 transfer-to-me 27-Jul | 27-Jul: +0.00 selftest neutral | 28-Jul: -130.00 transfer-to-me 28-Jul | 28-Jul: +396.00 owner-paid 28-Jul | 29-Jul: -205.00 transfer-to-me 29-Jul | 29-Jul: +60.00 owner-paid 29-Jul | 30-Jul: +268.00 owner-paid 30-Jul | 31-Jul: -390.00 transfer-to-me 31-Jul | 31-Jul: +78.00 owner-paid 31-Jul | 03-Aug: -225.00 transfer-to-me 01-Aug
    commission_rate=0.0406,             # Mercado Pago est. on card revenue
    soft_commission_rate=0.0205,        # Soft Restaurant terminal (reference only; actual value stored per-day in col V)
    bbva_commission_rate=0.0190,        # BBVA terminal (reference only; actual value stored per-day in col Z)
    week1_start=date(2026,5,18),        # W1 begins here; weeks are 7-day blocks
)

def _money(n): return f"${n:,.0f}"
def _signed(n): return (f"+${n:,.0f}" if n>=0 else f"-${abs(n):,.0f}")

def _gather():
    """Full operating dataset from Daily Log + Expenses (raw values, no formulas)."""
    from collections import defaultdict
    wb = openpyxl.load_workbook(P)
    dl = wb.worksheets[S_DAILY]; ex = wb.worksheets[S_EXP]
    last = _last_data_row(ex)
    ebd = defaultdict(float); cat_m = defaultdict(lambda: defaultdict(float)); capf_bd = defaultdict(float)
    for r in range(8, last+1):
        v = ex.cell(r,2).value
        if not isinstance(v,(datetime,date)): continue
        dd = v.date() if isinstance(v,datetime) else v
        amt = float(ex.cell(r,6).value or 0)
        ebd[dd]+=amt; cat_m[(dd.year,dd.month)][str(ex.cell(r,5).value or 'Other')]+=amt
        if str(ex.cell(r,7).value or '')=='Capital': capf_bd[dd]+=amt   # paid from Capital: excluded from operating-cash roll
    days=[]; inc_m=defaultdict(lambda: defaultdict(float)); dcount=defaultdict(int); soft_bd=defaultdict(float); softcomm_bd=defaultdict(float); bbva_bd=defaultdict(float); bbvacomm_bd=defaultdict(float)
    for r in range(2, dl.max_row+1):
        v = dl.cell(r,2).value
        if v is None: continue
        dd = v.date() if isinstance(v,datetime) else v
        if not isinstance(dd,date): continue
        ca=float(dl.cell(r,4).value or 0); cs=float(dl.cell(r,5).value or 0); tf=float(dl.cell(r,7).value or 0)
        sf=float(dl.cell(r,21).value or 0)   # col U = Soft Restaurant card revenue
        if sf: soft_bd[dd]+=sf
        vc=dl.cell(r,22).value               # col V = Soft commission (actual value, user-provided)
        if isinstance(vc,(int,float)): softcomm_bd[dd]+=float(vc)
        bb=float(dl.cell(r,25).value or 0)   # col Y = BBVA card revenue
        if bb: bbva_bd[dd]+=bb
        zc=dl.cell(r,26).value               # col Z = BBVA commission (actual value, user-provided)
        if isinstance(zc,(int,float)): bbvacomm_bd[dd]+=float(zc)
        days.append((dd,ca,cs,tf,round(ebd.get(dd,0.0),2)))
        k=(dd.year,dd.month); inc_m[k]['Card']+=ca+sf+bb; inc_m[k]['Cash']+=cs; inc_m[k]['Transfer']+=tf
        if ca+cs+tf+sf+bb>0: dcount[k]+=1
    days.sort()
    # --- closed-day expenses (Sundays) must still reduce cash -------------------
    # Sundays have expenses (the weekly shop) but NO Daily Log row, so their spend
    # was never subtracted in the cash roll -> Cash on Hand came out overstated.
    # Add a zero-revenue synthetic day for any expense date with no Daily Log row.
    _dlseen = {d for d,_,_,_,_ in days}
    orphan_dates = sorted(d for d in ebd if d not in _dlseen and ebd.get(d,0.0))
    for d in orphan_dates:
        days.append((d, 0.0, 0.0, 0.0, round(ebd.get(d,0.0),2)))
    days.sort()
    # owner ledger balance (raw sums)
    ol = wb[OWNER]; spent=transferred=0.0
    for r in range(4, ol.max_row+1):
        if str(ol.cell(r,1).value or '').startswith('TOTALS'): break
        spent+=float(ol.cell(r,4).value or 0); transferred+=float(ol.cell(r,5).value or 0)
    # advance still owed by Operations to Capital (Capital & Ownership, Section K, C147)
    try:
        owe_capital = round(float(wb.worksheets[0]['C147'].value or 0), 2)
    except Exception:
        owe_capital = 0.0
    # ---- partner capital positions (for the 3 ownership cards on the dashboard) ----
    # These were hard-coded and went stale (Shashi showed "owes $1,105" long after a
    # -$1,034 reimbursement had netted her to exactly her $120k budget).
    partners = {}
    try:
        cp = wb.worksheets[0]
        def _s(lo, hi):
            return sum(float(cp.cell(r,4).value or 0) for r in range(lo,hi+1)
                       if isinstance(cp.cell(r,4).value,(int,float)))
        lo_b = float(cp.cell(6,4).value or 0)      # 300,000
        ka_b = float(cp.cell(7,4).value or 0)      # 180,000
        sh_b = float(cp.cell(8,4).value or 0)      # 120,000
        ka_dep = float(cp.cell(42,4).value or 0) + _s(43,51)
        sh_dep = float(cp.cell(71,4).value or 0) + _s(72,81)
        # Lohith holds his own budget + what the other two transferred TO him
        funds_in = lo_b + float(cp.cell(42,4).value or 0) + float(cp.cell(71,4).value or 0)
        lo_remaining = funds_in - (_s(56,66) - float(cp.cell(56,4).value or 0)) \
                       - _s(14,30) - owe_capital
        partners = {
            'Lohith Reddy':   dict(dep=round(funds_in-lo_remaining,2), other=round(lo_remaining,2),
                                   base=funds_in),
            'Kashigoud Patil':dict(dep=round(ka_dep,2), other=round(ka_b-ka_dep,2), base=ka_b),
            'Shashirekha B.': dict(dep=round(sh_dep,2), other=round(sh_b-sh_dep,2), base=sh_b),
        }
    except Exception:
        partners = {}
    return dict(days=days, ebd=ebd, capf_bd=capf_bd, soft_bd=soft_bd, softcomm_bd=softcomm_bd, bbva_bd=bbva_bd, bbvacomm_bd=bbvacomm_bd, cat_m=cat_m, inc_m=inc_m, dcount=dcount,
                ol_spent=round(spent,2), ol_transferred=round(transferred,2), ol_balance=round(spent-transferred,2),
                owe_capital=owe_capital, partners=partners, orphan_dates=orphan_dates)


def _build_bars(days, xbd=None):
    xbd = xbd or {}
    recent = days[-30:]
    maxrev = max((c+k+t+xbd.get(d,0.0) for d,c,k,t,_ in recent), default=1) or 1
    out=[]
    for d,c,k,t,e in recent:
        rev=c+k+t+xbd.get(d,0.0); net=rev-e
        widx=(d-CFG['week1_start']).days//7
        col=WEEK_COLORS[widx % len(WEEK_COLORS)]
        w=max(6, round(rev/maxrev*100)) if rev>0 else 3
        cls='up' if net>=0 else 'down'
        out.append(f'<div class="bar-row"><div class="bar-day">{d.strftime("%d-%b")}</div>'
                   f'<div class="bar-track"><div class="bar-fill" style="width:{w}%;background:{col}">{_money(rev)}</div></div>'
                   f'<div class="bar-amt">{_money(rev)}</div><div class="bar-net {cls}">{_signed(net)}</div></div>')
    return ''.join(out)

def _build_weeks(days, xbd=None):
    xbd = xbd or {}
    from collections import defaultdict
    wk=defaultdict(list)
    for row in days: wk[(row[0]-CFG['week1_start']).days//7].append(row)
    idxs=sorted(wk)[-8:]
    maxrev=max((sum(c+k+t+xbd.get(d,0.0) for d,c,k,t,_ in wk[i]) for i in idxs), default=1) or 1
    out=[]
    for i in idxs:
        rows=wk[i]; rev=sum(c+k+t+xbd.get(d,0.0) for d,c,k,t,_ in rows); exp=sum(e for *_,e in rows); net=rev-exp
        ds=[r[0] for r in rows]; lo,hi=min(ds),max(ds)
        if lo.month==hi.month: rng=f'{lo.day}-{hi.day} {lo.strftime("%b")}'
        else: rng=f'{lo.day} {lo.strftime("%b")}-{hi.day} {hi.strftime("%b")}'
        col=WEEK_COLORS[i % len(WEEK_COLORS)]
        w=max(4, round(rev/maxrev*100))
        cls='up' if net>=0 else 'down'
        out.append(f'<div class="wk"><div class="wk-label">W{i+1} {rng}</div>'
                   f'<div class="wk-rev" style="color:{col}">{_money(rev)}</div>'
                   f'<div style="background:var(--border);height:3px;border-radius:2px;margin:5px 0;overflow:hidden;"><div style="width:{w}%;height:100%;background:{col};"></div></div>'
                   f'<div class="wk-row"><span>Expenses</span><span>{_money(exp)}</span></div>'
                   f'<div class="wk-row"><span>Net</span><span class="{cls}">{_signed(net)}</span></div></div>')
    return ''.join(out)

GROUP_ORDER = ['Salaries','Groceries','Supplies','Utilities','Rent','Maintenance','Software','Other']
def _cat_group(c):
    c=str(c or '')
    if c.startswith('Staff'): return 'Salaries'
    if c.startswith('Ingredients') or c.startswith('Supermarket'): return 'Groceries'
    if c.startswith('Utilities'): return 'Utilities'
    if c=='Rent': return 'Rent'
    if c=='Maintenance': return 'Maintenance'
    if c.startswith('Software'): return 'Software'
    if ('Supplies' in c) or ('Disposables' in c) or ('Packaging' in c): return 'Supplies'
    return 'Other'

def _build_monthly(inc_m, cat_m, dcount):
    from collections import defaultdict
    months=sorted(set(list(inc_m)+list(cat_m))); CUR=months[-1]
    lab={m: date(m[0],m[1],1).strftime('%b %Y') for m in months}
    ordered=list(reversed(months))[:6]
    cards=[]
    for m in ordered:
        ti=sum(inc_m[m].values()); te=sum(cat_m[m].values()); net=ti-te; cur=(m==CUR)
        nc='var(--green)' if net>=0 else 'var(--red)'; bd='var(--green)' if cur else 'var(--border)'
        bg='linear-gradient(135deg,#0d2818,#161b22)' if cur else '#161b22'
        bdg='<span style="font-size:.55rem;background:var(--green);color:#04140a;padding:1px 6px;border-radius:8px;font-weight:800;margin-left:6px;vertical-align:middle;">LIVE</span>' if cur else ''
        cards.append(f'<div style="flex:1;min-width:175px;border:1px solid {bd};border-radius:10px;padding:13px 15px;background:{bg};">'
            f'<div style="font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-bottom:8px;">{lab[m]}{bdg}</div>'
            f'<div style="display:flex;justify-content:space-between;font-size:.78rem;margin:3px 0;"><span style="color:var(--muted);">Income</span><span style="color:var(--green);font-weight:700;">{_money(ti)}</span></div>'
            f'<div style="display:flex;justify-content:space-between;font-size:.78rem;margin:3px 0;"><span style="color:var(--muted);">Expenses</span><span style="color:var(--red);font-weight:700;">{_money(te)}</span></div>'
            f'<div style="display:flex;justify-content:space-between;font-size:.9rem;margin:6px 0 2px;border-top:1px solid var(--border);padding-top:6px;"><span style="font-weight:700;">Net</span><span style="color:{nc};font-weight:800;">{_signed(net)}</span></div>'
            f'<div style="font-size:.62rem;color:var(--muted);margin-top:4px;">{dcount[m]} trading days</div></div>')
    def tr(lbl,vals,color=None,bold=False,hdr=False,indent=False):
        td=''
        for v in vals:
            st=('font-weight:700;' if bold else '')+(f'color:{color};' if color else '')
            td+=f'<td style="text-align:right;padding:4px 10px;{st}border-bottom:1px solid var(--border);">{v}</td>'
        ls='font-weight:700;' if (bold or hdr) else ''
        pl='padding-left:26px;color:var(--muted);font-size:.92em;' if indent else ''
        return f'<tr><td style="text-align:left;padding:4px 10px;{ls}{pl}border-bottom:1px solid var(--border);">{lbl}</td>{td}</tr>'
    th='<tr><th style="text-align:left;padding:5px 10px;color:var(--muted);font-size:.66rem;text-transform:uppercase;">Line</th>'+''.join(f'<th style="text-align:right;padding:5px 10px;color:var(--muted);font-size:.66rem;text-transform:uppercase;">{lab[m]}</th>' for m in ordered)+'</tr>'
    grp_m={m:defaultdict(float) for m in ordered}
    for m in ordered:
        for c,a in cat_m[m].items(): grp_m[m][_cat_group(c)]+=a
    gtot=lambda g: sum(grp_m[m].get(g,0) for m in ordered)
    groups=[g for g in GROUP_ORDER if gtot(g)>0]
    # current-month expense pie (donut)
    GCOL={'Salaries':'#3b82f6','Groceries':'#22c55e','Supplies':'#f97316','Utilities':'#a855f7','Rent':'#ef4444','Maintenance':'#f59e0b','Software':'#06b6d4','Other':'#64748b'}
    cm=ordered[0]; ctot=sum(grp_m[cm].values()); pie=''
    if ctot>0:
        segs=[]; leg=[]; off=0.0
        for g in groups:
            v=grp_m[cm].get(g,0)
            if v<=0: continue
            pct=v/ctot*100; col=GCOL.get(g,'#64748b')
            segs.append(f'<circle cx="18" cy="18" r="15.915" fill="none" stroke="{col}" stroke-width="4.4" stroke-dasharray="{pct:.3f} {100-pct:.3f}" stroke-dashoffset="{25-off:.3f}"/>')
            leg.append(f'<div style="display:flex;align-items:center;gap:7px;font-size:.72rem;margin:3px 0;"><span style="width:10px;height:10px;border-radius:2px;background:{col};flex:none;"></span><span style="flex:1;">{g}</span><span style="font-weight:700;">{_money(v)}</span><span style="color:var(--muted);width:40px;text-align:right;">{pct:.0f}%</span></div>')
            off+=pct
        donut=(f'<svg viewBox="0 0 36 36" style="width:128px;height:128px;flex:none;">'
               f'<circle cx="18" cy="18" r="15.915" fill="none" stroke="var(--border)" stroke-width="4.4"/>{"".join(segs)}'
               f'<text x="18" y="16.6" text-anchor="middle" style="font-size:3px;fill:var(--muted);">{lab[cm]}</text>'
               f'<text x="18" y="21.2" text-anchor="middle" style="font-size:4.2px;font-weight:700;fill:#e6edf3;">{_money(ctot)}</text></svg>')
        pie=(f'<div style="display:flex;gap:20px;align-items:center;flex-wrap:wrap;margin-top:16px;padding:15px;border:1px solid var(--border);border-radius:10px;background:#0e141b;">'
             f'<div style="display:flex;flex-direction:column;align-items:center;">{donut}<div style="font-size:.6rem;color:var(--muted);margin-top:5px;">expense split</div></div>'
             f'<div style="flex:1;min-width:210px;"><div style="font-size:.7rem;text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-bottom:7px;">{lab[cm]} &mdash; expenses by group</div>{"".join(leg)}</div></div>')
    body=tr('INCOME',['' for _ in ordered],hdr=True)
    for typ in ['Card','Cash','Transfer']: body+=tr(typ,[_money(inc_m[m].get(typ,0)) for m in ordered])
    body+=tr('Total Income',[_money(sum(inc_m[m].values())) for m in ordered],color='var(--green)',bold=True)
    body+=tr('EXPENSES BY GROUP',['' for _ in ordered],hdr=True)
    for g in groups: body+=tr(g,[_money(grp_m[m].get(g,0)) for m in ordered])
    body+=tr('Total Expenses',[_money(sum(cat_m[m].values())) for m in ordered],color='var(--red)',bold=True)
    body+=tr('NET PROFIT / LOSS',[_signed(sum(inc_m[m].values())-sum(cat_m[m].values())) for m in ordered],bold=True)
    table=f'<div style="overflow-x:auto;margin-top:14px;"><table style="width:100%;border-collapse:collapse;font-size:.74rem;"><thead>{th}</thead><tbody>{body}</tbody></table></div>'
    dbody=''
    for g in groups:
        cs=sorted({c for m in ordered for c in cat_m[m] if _cat_group(c)==g}, key=lambda c:-sum(cat_m[m].get(c,0) for m in ordered))
        dbody+=tr(g,[_money(grp_m[m].get(g,0)) for m in ordered],bold=True)
        for c in cs: dbody+=tr(c,[_money(cat_m[m].get(c,0)) for m in ordered],indent=True)
    dtable=f'<div style="overflow-x:auto;margin-top:10px;"><table style="width:100%;border-collapse:collapse;font-size:.72rem;"><thead>{th}</thead><tbody>{dbody}</tbody></table></div>'
    details=f'<details style="margin-top:12px;"><summary style="cursor:pointer;font-size:.72rem;color:var(--accent,#3b82f6);user-select:none;">&#9656; Show detailed categories</summary>{dtable}</details>'
    note='<p style="font-size:.66rem;color:var(--muted);margin-top:10px;line-height:1.5;">Expenses grouped into Salaries, Groceries, Supplies, Utilities, Rent, Maintenance &amp; Software. Rent + most salaries were front-loaded into June, so June reads heavier; May (opening fortnight) shows a large profit. Expand &ldquo;detailed categories&rdquo; for the full per-item list.</p>'
    return (f'<div class="card" style="margin-bottom:16px;"><h3>&#128197; Monthly Breakdown</h3>'
            f'<div style="display:flex;gap:12px;flex-wrap:wrap;">{"".join(cards)}</div>{pie}{table}{details}{note}</div>')

def refresh_dashboard(do_backup=True):
    import re
    g = _gather(); days = g['days']
    soft_bd=g['soft_bd']; softcomm_bd=g['softcomm_bd']; soft_total=round(sum(soft_bd.values()),2)
    soft_comm=round(sum(softcomm_bd.values()),2)   # actual Soft Restaurant commission (user-provided per day)
    bbva_bd=g['bbva_bd']; bbvacomm_bd=g['bbvacomm_bd']; bbva_total=round(sum(bbva_bd.values()),2)
    bbva_comm=round(sum(bbvacomm_bd.values()),2)   # actual BBVA commission (user-provided per day)
    rev=sum(c+k+t for _,c,k,t,_ in days)+soft_total+bbva_total; exp=round(sum(g['ebd'].values()),2)
    net=round(rev-exp,2)
    # Operating Net = trading result, adding back only the operating costs STILL funded by
    # Capital (Section K C147). Was hard-coded +20,000 (the Jul-5 rent), but Operations has
    # since repaid $22,587 of the advance out of trading cash -- once repaid, the business
    # HAS borne that cost, so adding the full amount back overstated the result.
    # Deriving it from C147 means every repayment lowers this automatically.
    netrent=round(net + g['owe_capital'], 2)
    trading=sum(1 for _,c,k,t,_ in days if c+k+t>0)
    card=sum(d[1] for d in days)
    rate=CFG['commission_rate']
    comm=round(card*rate + soft_comm + bbva_comm)   # MP (4.06% auto) + Soft Restaurant + BBVA (actual per-day)
    last_d,lc,lk,lt,le = days[-1]
    last_rev=lc+lk+lt+soft_bd.get(last_d,0.0)+bbva_bd.get(last_d,0.0); last_net=last_rev-le
    # cash on hand = anchor + nets after anchor + soft/bbva card net - commissions after anchor + manual adjust
    roll=sum((c+k+t-(e-g['capf_bd'].get(d,0.0))) for d,c,k,t,e in days if d>CFG['cash_anchor_date'])
    soft_post=round(sum(v for dd,v in soft_bd.items() if dd>CFG['cash_anchor_date']),2)
    softcomm_post=round(sum(v for dd,v in softcomm_bd.items() if dd>CFG['cash_anchor_date']),2)
    bbva_post=round(sum(v for dd,v in bbva_bd.items() if dd>CFG['cash_anchor_date']),2)
    bbvacomm_post=round(sum(v for dd,v in bbvacomm_bd.items() if dd>CFG['cash_anchor_date']),2)
    comm_post=sum(c*rate for d,c,k,t,e in days if d>CFG['cash_anchor_date'])+softcomm_post+bbvacomm_post
    cash=round(CFG['cash_anchor_amount']+roll+soft_post+bbva_post-comm_post+CFG['cash_adjust'])
    net_allin=round(rev-exp-comm,2)       # bottom line after ALL expenses AND commission
    datestr=last_d.strftime('%a %d %b %Y')
    if do_backup:
        import shutil, os, glob as _glob
        bdir=r'C:\LOKA\Backup'; os.makedirs(bdir,exist_ok=True)
        shutil.copy(DASH, os.path.join(bdir, 'dashboard_'+datetime.now().strftime('%Y%m%d_%H%M%S')+'.html'))
        # rotate: keep only the newest 15 dashboard snapshots (they used to pile up to 100+)
        snaps=sorted(_glob.glob(os.path.join(bdir,'dashboard_*.html')), key=os.path.getmtime, reverse=True)
        for old in snaps[15:]:
            try: os.remove(old)
            except OSError: pass
    s=open(DASH,encoding='utf-8').read()
    def sub(pat, repl, label, n=None):
        nonlocal s
        s2,cnt=re.subn(pat, repl, s)
        if n is not None and cnt!=n: print(f'  WARN {label}: expected {n} got {cnt}')
        elif cnt==0: print(f'  WARN {label}: no match')
        s=s2
    M=lambda x: _money(x).replace('$','\\$')  # not used; placeholder
    # --- scalars (label-anchored) ---
    sub(r'Updated: [^<]+', f'Updated: {datestr} (EOD)', 'header')
    sub(r'(<div class="bval">)\$[\d,]+(</div>)', lambda m: m.group(1)+_money(cash)+m.group(2), 'cash')
    sub(r'(All-Time Revenue</div><div[^>]*>)\$[\d,]+', lambda m: m.group(1)+_money(rev), 'banner rev')
    sub(r'(All-Time Expenses</div><div[^>]*>)\$[\d,]+', lambda m: m.group(1)+_money(exp), 'banner exp')
    sub(r'(Commission</div><div[^>]*>)\$[\d,]+', lambda m: m.group(1)+_money(comm), 'banner comm')
    sub(r'(<div class="klabel">Total Revenue</div><div class="kval">)\$[\d,]+(</div><div[^>]*>)\d+( trading days)',
        lambda m: m.group(1)+_money(rev)+m.group(2)+str(trading)+m.group(3), 'kpi rev')
    sub(r'(<div class="klabel">Total Expenses</div><div class="kval">)\$[\d,]+', lambda m: m.group(1)+_money(exp), 'kpi exp')
    sub(r'(<div class="klabel">Operating Net</div><div class="kval">)[+\-]?\$[\d,]+', lambda m: m.group(1)+_signed(netrent), 'kpi net')
    # keep the little grey descriptor under Operating Net honest about WHAT is excluded
    sub(r'(<div class="klabel">Operating Net</div><div class="kval">[+\-]?\$[\d,]+</div><div style="font-size:\.66rem;color:var\(--muted\);">)[^<]*',
        lambda m: m.group(1)+f'excl. {_money(g["owe_capital"])} still owed to Capital', 'kpi net note')
    sub(r'(MP Commission</div><div class="kval">)\$[\d,]+', lambda m: m.group(1)+_money(comm), 'kpi comm')
    sub(r'(<div class="alabel">)[^<]+(</div><div class="aval">)Rev [^<]+',
        lambda m: m.group(1)+last_d.strftime('%a %d %b')+m.group(2)+f'Rev {_money(last_rev)} &mdash; Exp {_money(le)} &mdash; Net {_signed(last_net)}', 'daily alert')
    sub(r'Dashboard updated [^<]+',
        f'Dashboard updated {datestr} (EOD) &bull; auto-refreshed by loka.py', 'footer')
    # owner ledger card + alert
    sub(r'(Spent personally</span><span class="sval" style="color:var\(--red\);">)\$[\d,]+', lambda m: m.group(1)+_money(g['ol_spent']), 'ledger spent')
    sub(r'(Transfers received</span><span class="sval" style="color:var\(--green\);">)\$[\d,]+', lambda m: m.group(1)+_money(g['ol_transferred']), 'ledger recv')
    # ledger balance card + alert: the sign FLIPS (positive = restaurant owes Lohith,
    # negative = Lohith is holding restaurant money), so both the label and the value
    # must be rewritten, and the value shown as an absolute amount.
    _olb = g['ol_balance']
    _ollab = 'Restaurant owes Lohith' if _olb > 0 else ('Lohith holds' if _olb < 0 else 'Ledger settled')
    sub(r'(<span class="slabel" style="color:var\(--accent\);">)(?:Restaurant owes Lohith|Lohith holds|Ledger settled)(</span><span class="sval"[^>]*>)\$-?[\d,]+',
        lambda m: m.group(1)+_ollab+m.group(2)+_money(abs(_olb)), 'ledger bal')
    sub(r'(Lohith Ledger</div><div class="aval">)(?:Restaurant owes|Lohith holds|Settled) \$-?[\d,]+',
        lambda m: m.group(1)+('Restaurant owes' if _olb>0 else ('Lohith holds' if _olb<0 else 'Settled'))+' '+_money(abs(_olb)), 'ledger alert')
    # operations <-> capital (Section K) — commission + true bottom line
    _s3=lambda n: (f'+${n:,.0f}' if n>=0 else f'&minus;${abs(n):,.0f}')
    sub(r'(Card commissions \(MP\+BBVA\+Soft\)</span><span class="sval"[^>]*>)(?:&minus;)?\$[\d,]+', lambda m: m.group(1)+'&minus;'+_money(comm), 'ops comm')
    def _netitem(m):
        pos=net_allin>=0; col='var(--green)' if pos else 'var(--red)'; bg='#0f2a12' if pos else '#2a0f0f'
        return (f'<div class="stat-item" style="background:{bg};padding:6px 8px;border-radius:6px;margin-top:5px;">'
                f'<span class="slabel" style="color:{col};">Net after all exp &amp; commission</span>'
                f'<span class="sval" style="color:{col};font-size:.9rem;">{_s3(net_allin)}</span></div>')
    sub(r'<div class="stat-item" style="background:#[0-9a-fA-F]{6};padding:6px 8px;border-radius:6px;margin-top:5px;"><span class="slabel" style="color:var\(--(?:red|green)\);">Net after all exp &amp; commission</span><span class="sval" style="color:var\(--(?:red|green)\);font-size:.9rem;">(?:\+|&minus;)?\$[\d,]+</span></div>', _netitem, 'ops net-allin')
    sub(r'net after all exp &amp; comm (?:\+|&minus;)?\$[\d,]+', lambda m: 'net after all exp &amp; comm '+_s3(net_allin), 'banner net-allin')
    # ---- partner ownership cards: were HARD-CODED and went stale (Shashi still showed
    # "owes $1,105" long after a -$1,034 reimbursement netted her to her exact $120k).
    # Now driven straight off the Capital sheet.
    def _owncard(m):
        card = m.group(0)
        nm = re.search(r'<div class="oname">([^<]+)</div>', card)
        if not nm or nm.group(1) not in g.get('partners', {}): return card
        p = g['partners'][nm.group(1)]
        pct = max(0.0, min(100.0, (p['dep']/p['base']*100) if p['base'] else 0))
        card = re.sub(r'(<div class="own-bar" style="width:)[\d.]+(%)',
                      lambda x: x.group(1)+f'{pct:.1f}'+x.group(2), card)
        it = iter([_money(p['dep']), _money(abs(p['other']))])   # row1 deployed, row2 owes/remaining
        card = re.sub(r'(<div class="ov"(?: style="[^"]*")?>)\$[\d,]+(</div>)',
                      lambda x: x.group(1)+next(it, x.group(0)[len(x.group(1)):-len(x.group(2))])+x.group(2), card)
        return card
    sub(r'<div class="own-card">.*?</div></div></div>', _owncard, 'partner cards', 3)
    # --- net CASH POSITION (balance-sheet view): cash held minus what operations owes ---
    net_cash_pos = round(cash - g['owe_capital'] - g['ol_balance'], 2)
    sub(r'Ops owe capital \$[\d,]+', lambda m: 'Ops owe capital '+_money(g['owe_capital']), 'banner ops-owe')
    def _ncp(m):
        pos = net_cash_pos>=0; col='var(--green)' if pos else 'var(--red)'
        return m.group(1)+f'<div style="font-size:.85rem;font-weight:700;color:{col}">'+_s3(net_cash_pos)+'</div>'
    sub(r'(Net Cash Position</div>)<div[^>]*>(?:\+|&minus;)?\$[\d,]+</div>', _ncp, 'banner net-cash-pos')
    sub(r'(Net Cash Position[^<]*</span><span class="sval"[^>]*>)(?:\+|&minus;)?\$[\d,]+',
        lambda m: m.group(1)+_s3(net_cash_pos), 'ops net-cash-pos')
    # --- block regens ---
    xbd={d: soft_bd.get(d,0.0)+bbva_bd.get(d,0.0) for d in set(soft_bd)|set(bbva_bd)}  # extra card revenue (Soft+BBVA) per day
    bars=_build_bars(days, xbd)
    s=re.sub(r'(<h3>Daily Revenue &mdash; Last 30 Days</h3>).*?(<div style="display:flex;gap:14px;margin-top:10px;)',
             lambda m: m.group(1)+bars+m.group(2), s, count=1, flags=re.DOTALL)
    weeks=_build_weeks(days, xbd)
    s=re.sub(r'(<div class="week-grid">).*?(</div>\s*<p class="note")',
             lambda m: m.group(1)+weeks+m.group(2), s, count=1, flags=re.DOTALL)
    monthly=_build_monthly(g['inc_m'], g['cat_m'], g['dcount'])
    s=re.sub(r'<div class="card" style="margin-bottom:16px;"><h3>&#128197; Monthly Breakdown</h3>.*?(<h3 style="margin-bottom:10px;">&#128202; All-Time)',
             lambda m: monthly+m.group(1), s, count=1, flags=re.DOTALL)
    open(DASH,'w',encoding='utf-8',newline='\n').write(s)
    # keep the Monthly P&L sheet's structure current too (formulas auto-update values themselves)
    try:
        info=refresh_pl(do_backup=False); print(f'  Monthly P&L synced: {info["months"]} months, {info["cats"]} categories')
    except Exception as e:
        print(f'  (Monthly P&L sync skipped: {e})')
    print(f'Dashboard refreshed @ {datestr}: rev {_money(rev)} exp {_money(exp)} opnet {_signed(netrent)} | cash {_money(cash)} | comm {_money(comm)} | ledger {_money(g["ol_balance"])}')


# ============== MONTHLY P&L (LIVE FORMULAS) ==============
def refresh_pl(do_backup=True):
    """Rebuild the Monthly P&L sheet with LIVE SUMIFS/COUNTIFS formulas that
    pull from Daily Log + Expenses, so it auto-updates whenever data changes."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    if do_backup: backup('monthly_pl')
    wb = openpyxl.load_workbook(P)
    dl = wb.worksheets[S_DAILY]; ex = wb.worksheets[S_EXP]; pl = wb.worksheets[8]
    DLN = "'" + dl.title + "'"; EXN = "'" + ex.title + "'"
    # detect months (from both sheets) and categories (from expenses), in data order
    def parse(v):
        from datetime import datetime as _dt
        if isinstance(v,_dt): return v.date()
        if isinstance(v,date): return v
        return None
    monthset=set(); cat_tot=defaultdict(float)
    for r in range(2, dl.max_row+1):
        d=parse(dl.cell(r,2).value)
        if d: monthset.add((d.year,d.month))
    last=ex.max_row
    while last>7 and ex.cell(last,3).value is None: last-=1
    for r in range(8,last+1):
        d=parse(ex.cell(r,2).value)
        if d:
            monthset.add((d.year,d.month))
            cat_tot[str(ex.cell(r,5).value or 'Other')]+=float(ex.cell(r,6).value or 0)
    months=sorted(monthset)
    cats=[c for c,_ in sorted(cat_tot.items(), key=lambda x:-x[1])]
    label={m: date(m[0],m[1],1).strftime('%b %Y') for m in months}
    ncol=1+len(months)
    def nextm(m):
        y,mo=m; return (y+1,1) if mo==12 else (y,mo+1)
    # date-range criteria builder for a sheet/col
    def rng(sheet,col,m):
        y,mo=m; ny,nm=nextm(m)
        return (f'{sheet}!${col}$8:${col}$5000,{sheet}!$B$8:$B$5000,">="&DATE({y},{mo},1),'
                f'{sheet}!$B$8:$B$5000,"<"&DATE({ny},{nm},1)')
    return _write_pl(wb, pl, months, cats, label, ncol, DLN, EXN, rng, P)


def _write_pl(wb, pl, months, cats, label, ncol, DLN, EXN, rng, Ppath):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    # strip
    for mr in list(pl.merged_cells.ranges): pl.unmerge_cells(str(mr))
    for r in range(1, max(pl.max_row,60)+1):
        for c in range(1, max(pl.max_column,12)+1):
            cell=pl.cell(r,c); cell.value=None
            cell.font=Font(); cell.fill=PatternFill(fill_type=None); cell.border=Border()
            cell.alignment=Alignment(); cell.number_format='General'
    for col in 'ABCDEFGHIJKLMN': pl.column_dimensions[col].width=9.0
    TITLE=Font(bold=True,color='FFFFFF',size=13); TFILL=PatternFill('solid',fgColor='0D2818')
    HDR=Font(bold=True,color='FFFFFF',size=10); HFILL=PatternFill('solid',fgColor='1F4E2C')
    BOLD=Font(bold=True); NORM=Font()
    thin=Side(style='thin',color='D9D9D9'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
    MONEY='$#,##0.00'; INTF='#,##0'
    RIGHT=Alignment(horizontal='right'); LEFT=Alignment(horizontal='left'); CTR=Alignment(horizontal='center')
    def cel(r,c,val,font=NORM,fill=None,fmt=None,align=None):
        x=pl.cell(r,c,val); x.font=font; x.border=BORD
        if fill: x.fill=fill
        if fmt: x.number_format=fmt
        x.alignment=align or (RIGHT if c>1 else LEFT)
    def hrow(r,text):
        cel(r,1,text,font=HDR,fill=HFILL,align=LEFT)
        for j,m in enumerate(months): cel(r,2+j,label[m],font=HDR,fill=HFILL,align=RIGHT)
    # title
    pl.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    t=pl.cell(1,1,'LOKA — MONTHLY P&L   (live formulas — auto-updates from Daily Log & Expenses)')
    t.font=TITLE; t.fill=TFILL; t.alignment=CTR
    for c in range(1,ncol+1): pl.cell(1,c).fill=TFILL
    pl.row_dimensions[1].height=26
    COL=lambda j: get_column_letter(2+j)
    r=3
    hrow(r,'SUMMARY'); r+=1
    R_INC=r; cel(r,1,'Total Income',font=BOLD)
    for j,m in enumerate(months): cel(r,2+j,f'=SUMIFS({rng(DLN,"F",m)})',fmt=MONEY)
    r+=1
    R_EXP=r; cel(r,1,'Total Expenses',font=BOLD)
    for j,m in enumerate(months): cel(r,2+j,f'=SUMIFS({rng(EXN,"F",m)})',fmt=MONEY)
    r+=1
    cel(r,1,'NET PROFIT / LOSS',font=BOLD)
    for j in range(len(months)): cel(r,2+j,f'={COL(j)}{R_INC}-{COL(j)}{R_EXP}',font=BOLD,fmt=MONEY)
    r+=1
    R_DAYS=r; cel(r,1,'Trading Days')
    for j,m in enumerate(months): cel(r,2+j,f'=COUNTIFS({DLN}!$F$8:$F$5000,">0",'+rng(DLN,"F",m).split(",",1)[1]+')',fmt=INTF)
    r+=1
    cel(r,1,'Avg Income / Day')
    for j in range(len(months)): cel(r,2+j,f'=IFERROR({COL(j)}{R_INC}/{COL(j)}{R_DAYS},0)',fmt=MONEY)
    r+=2
    hrow(r,'INCOME BY TYPE'); r+=1
    R_CARD=r
    for col,name in [('D','Card'),('E','Cash'),('G','Transfer')]:
        cel(r,1,name)
        for j,m in enumerate(months): cel(r,2+j,f'=SUMIFS({rng(DLN,col,m)})',fmt=MONEY)
        r+=1
    cel(r,1,'Total Income',font=BOLD)
    for j in range(len(months)): cel(r,2+j,f'={COL(j)}{R_CARD}+{COL(j)}{R_CARD+1}+{COL(j)}{R_CARD+2}',font=BOLD,fmt=MONEY)
    r+=2
    hrow(r,'EXPENSES BY CATEGORY (high \u2192 low)'); r+=1
    for cat in cats:
        cel(r,1,cat)
        for j,m in enumerate(months):
            crit=f'{EXN}!$F$8:$F$5000,{EXN}!$E$8:$E$5000,$A{r},{EXN}!$B$8:$B$5000,">="&DATE({m[0]},{m[1]},1),{EXN}!$B$8:$B$5000,"<"&DATE({(m[0]+1) if m[1]==12 else m[0]},{1 if m[1]==12 else m[1]+1},1)'
            cel(r,2+j,f'=SUMIFS({crit})',fmt=MONEY)
        r+=1
    R_TEXP=r; cel(r,1,'TOTAL EXPENSES',font=BOLD)
    for j,m in enumerate(months): cel(r,2+j,f'=SUMIFS({rng(EXN,"F",m)})',font=BOLD,fmt=MONEY)
    r+=1
    cel(r,1,'NET PROFIT / LOSS',font=BOLD)
    for j in range(len(months)): cel(r,2+j,f'={COL(j)}{R_INC}-{COL(j)}{R_TEXP}',font=BOLD,fmt=MONEY)
    pl.column_dimensions['A'].width=34
    for j in range(len(months)): pl.column_dimensions[get_column_letter(2+j)].width=15
    pl.sheet_view.showGridLines=False
    try: wb.calculation.fullCalcOnLoad = True   # force Excel to recompute all formulas on open
    except Exception: pass
    wb.save(Ppath)
    return dict(months=len(months), cats=len(cats), rows=r)


if __name__ == '__main__':
    main()
