# -*- coding: utf-8 -*-
"""
LOKA - the jobs that used to need a Claude session.

  py tools.py recount        re-anchor cash to a physical count
  py tools.py repay          record money paid back to Capital
  py tools.py tips           record propinas (card tips handed to staff)
  py tools.py settle         settle the Owner Ledger with Lohith
  py tools.py report         weekly / monthly summary
  py tools.py restore        undo - roll back to a previous backup
"""
import sys, os, io, re, glob, shutil
from copy import copy
from datetime import datetime, date, timedelta
sys.path.insert(0, r'C:\LOKA')
import openpyxl, loka

P = loka.P
def ask(q, default=None):
    v = input(f"   {q}{' ['+str(default)+']' if default is not None else ''}: ").strip()
    return v or (str(default) if default is not None else '')
def money(q, default=None):
    while True:
        v = ask(q, default).replace(',','').replace('$','')
        try: return float(v)
        except ValueError: print("   Please type a number.")
def askdate(q='Date (e.g. 02-Aug or 02-Aug-2026)', default=None):
    default = default or date.today().strftime('%d-%b-%Y')
    while True:
        v = ask(q, default)
        try: return loka.pdate(v)
        except Exception: print("   Try a format like 02-Aug-2026.")
def confirm(msg):
    return ask(f"{msg} (y/n)", 'n').lower().startswith('y')

# ---------------------------------------------------------------- recount
def recount():
    print("\n== RE-ANCHOR CASH TO A PHYSICAL COUNT ==")
    print(f"   The books currently think you have: {loka.compute()['cash_on_hand']}")
    amt = money("Total cash you actually counted (all forms, incl bank)")
    d   = askdate("Date you counted it")
    print(f"\n   New anchor: {d:%d-%b-%Y} = ${amt:,.2f}")
    print("   This also resets the running cash adjustment to 0 (the count absorbs all drift).")
    if not confirm("   Apply?"): print("   cancelled."); return
    loka.backup(f'recount_{d:%b%d}'.lower())
    src = io.open(loka.__file__, encoding='utf-8').read()
    src = re.sub(r'cash_anchor_date=date\([^)]*\), cash_anchor_amount=[\d.]+',
                 f'cash_anchor_date=date({d.year},{d.month},{d.day}), cash_anchor_amount={amt}', src)
    src = re.sub(r'(cash_adjust=)-?[\d.]+(,\s*#)[^\n]*',
                 rf'\g<1>0.00\g<2> RESET at physical count {d:%d-%b-%Y} = ${amt:,.2f}. '
                 'Only add deltas dated AFTER that.', src)
    io.open(loka.__file__,'w',encoding='utf-8',newline='\n').write(src)
    print("   done - re-anchored.")
    os.system(f'"{sys.executable}" "{os.path.join(os.path.dirname(loka.__file__),"loka.py")}" refresh-all')

# ---------------------------------------------------------------- repay capital
def repay():
    print("\n== MONEY PAID BACK TO CAPITAL ==")
    wb = openpyxl.load_workbook(P); cap = wb.worksheets[0]
    owed = float(cap.cell(147,3).value or 0)
    print(f"   Operations currently owes Capital: ${owed:,.2f}")
    amt = money("How much was paid back")
    d   = askdate()
    if amt > owed and not confirm(f"   That is more than the ${owed:,.2f} owed. Continue?"): return
    new = round(owed-amt, 2)
    print(f"\n   Advance owed:  ${owed:,.2f}  ->  ${new:,.2f}")
    print(f"   Cash on hand will drop by ${amt:,.2f} (it left the restaurant).")
    print("   This is NOT a P&L expense - it is a balance-sheet movement.")
    if not confirm("   Apply?"): print("   cancelled."); return
    loka.backup(f'repay_capital_{d:%b%d}'.lower())
    cap.cell(147,3,new)
    prev = cap.cell(149,3).value or 0
    cap.cell(149,2,'  >> Repaid to Capital from operating cash')
    cap.cell(149,3, round(float(prev)-amt,2)); cap.cell(149,3).number_format='$#,##0.00'
    cap.cell(149,5, f'Cumulative repaid from operating cash (latest {d:%d-%b-%Y} ${amt:,.2f}). '
                    'Reduces advance owed & Cash-on-Hand; NOT a P&L expense.')
    wb.calculation.calcMode='auto'; wb.calculation.fullCalcOnLoad=True
    wb.save(P)
    loka.cash_adjust_add(-amt, f'capital repayment {d:%d-%b}')
    print(f"   done - advance now ${new:,.2f}")
    loka.refresh_all(do_backup=False)

# ---------------------------------------------------------------- propinas
def tips():
    print("\n== PROPINAS (card tips handed to staff) ==")
    print("   Reminder: tips are a PASS-THROUGH, not a restaurant expense.")
    print("   Only money the restaurant adds ON TOP of collected tips is a real cost.")
    wb = openpyxl.load_workbook(P); sp = wb['👥 Staff & Payroll']
    r = 63
    while r <= 92 and sp.cell(r,1).value not in (None,''): r += 1
    if r > 92:
        print("   The propinas table is FULL (rows 63-92). Ask Claude to extend it."); return
    print(f"   {92-r+1} free slot(s) left.")
    d = askdate()
    label = ask("Which week is this for (e.g. 'Week 11: 03-09 Aug')", f"week to {d:%d-%b}")
    entries = []
    for who in ('John','Duvi/Debi','Samu'):
        v = ask(f"Amount for {who} (blank/0 = skip)", '0').replace(',','').replace('$','')
        try: a = float(v)
        except ValueError: a = 0
        if a: entries.append((who, a))
    while confirm("   Add someone else?"):
        w = ask("Name"); a = money(f"Amount for {w}")
        if a: entries.append((w, a))
    if not entries: print("   nothing to record."); return
    print(f"\n   Will record {len(entries)} entr(y/ies), total ${sum(a for _,a in entries):,.2f}:")
    for w,a in entries: print(f"     {d:%d-%b}  {w} ({label})  ${a:,.2f}")
    if not confirm("   Apply?"): print("   cancelled."); return
    loka.backup(f'propinas_{d:%b%d}'.lower())
    tmpl = r-1 if r > 63 else 63
    for i,(w,a) in enumerate(entries):
        rr = r+i
        if rr > 92: print("   ran out of slots - stopped early."); break
        for c in (1,2,3,4):
            s_,dd = sp.cell(tmpl,c), sp.cell(rr,c)
            dd.font=copy(s_.font); dd.fill=copy(s_.fill); dd.border=copy(s_.border)
            dd.alignment=copy(s_.alignment); dd.number_format=s_.number_format
        sp.cell(rr,1,d); sp.cell(rr,1).number_format=sp.cell(tmpl,1).number_format
        sp.cell(rr,2,f'{w} ({label})'); sp.cell(rr,4,a)
    wb.calculation.calcMode='auto'; wb.calculation.fullCalcOnLoad=True
    wb.save(P)
    print(f"   done - recorded in rows {r}-{r+len(entries)-1}. Totals are NOT affected (pass-through).")

# ---------------------------------------------------------------- settle ledger
def settle():
    print("\n== SETTLE THE OWNER LEDGER ==")
    g = loka._gather(); bal = g['ol_balance']
    if abs(bal) < 0.01: print("   The ledger is already square."); return
    if bal > 0:
        print(f"   The restaurant owes you ${bal:,.2f}.")
        print("   Settling = the restaurant pays you -> cash goes DOWN.")
    else:
        print(f"   You are holding ${abs(bal):,.2f} of restaurant money.")
        print("   Settling = you hand it back -> cash goes UP.")
    amt = money("Amount settled", f"{abs(bal):.2f}")
    d   = askdate()
    if not confirm("   Apply?"): print("   cancelled."); return
    loka.backup(f'settle_ledger_{d:%b%d}'.lower())
    if bal > 0:
        loka.add_ledger(d, 'Settlement - restaurant paid Lohith', transferred=amt,
                        notes='Ledger settled from operating cash', do_backup=False)
        loka.cash_adjust_add(-amt, f'ledger settlement {d:%d-%b}')
    else:
        loka.add_ledger(d, 'Settlement - Lohith returned restaurant money', spent=amt,
                        notes='Lohith handed back cash he was holding', do_backup=False)
        loka.cash_adjust_add(amt, f'ledger settlement {d:%d-%b}')
    print("   done.")
    loka.refresh_all(do_backup=False)

# ---------------------------------------------------------------- report
def report():
    g = loka._gather()
    days = [(d,c,k,t,e) for d,c,k,t,e in g['days']]
    if not days: print("no data"); return
    last = max(d for d,_,_,_,_ in days)
    def block(title, sel):
        rows=[x for x in days if sel(x[0])]
        rev=sum(c+k+t for _,c,k,t,_ in rows)+sum(g['soft_bd'].get(d,0)+g['bbva_bd'].get(d,0) for d,_,_,_,_ in rows)
        exp=sum(e for *_,e in rows)
        trading=sum(1 for _,c,k,t,_ in rows if c+k+t>0)
        print(f"\n  {title}")
        print(f"    revenue {rev:>12,.2f}   expenses {exp:>12,.2f}   net {rev-exp:>12,.2f}")
        if trading: print(f"    {trading} trading day(s), avg revenue {rev/trading:,.2f}/day")
    print("\n" + "="*60); print(f"  LOKA REPORT  (latest data {last:%d-%b-%Y})"); print("="*60)
    wk = last - timedelta(days=last.weekday())
    block(f"THIS WEEK (from {wk:%d-%b})", lambda d: d >= wk)
    block(f"LAST 7 DAYS", lambda d: d > last - timedelta(days=7))
    block(f"THIS MONTH ({last:%B})", lambda d: d.month == last.month and d.year == last.year)
    block("ALL TIME", lambda d: True)
    s = loka.compute()['position']
    print(f"\n  cash on hand      ${s['cash_on_hand']:,.2f}")
    print(f"  owed to Capital   ${s['owed_to_capital']:,.2f}")
    print(f"  owner ledger      {'restaurant owes you' if s['owner_ledger']>0 else 'you hold'} ${abs(s['owner_ledger']):,.2f}")
    print(f"  net cash position ${s['net_cash_position']:,.2f}")
    print()

# ---------------------------------------------------------------- restore
def restore():
    print("\n== UNDO - RESTORE A PREVIOUS BACKUP ==")
    b = sorted(glob.glob(r'C:\LOKA\Backup\*.xlsx'), key=os.path.getmtime, reverse=True)[:15]
    if not b: print("   no backups found."); return
    for i,f in enumerate(b,1):
        t = datetime.fromtimestamp(os.path.getmtime(f))
        print(f"     {i:>2}. {t:%d-%b %H:%M}   {os.path.basename(f)[24:-5] or 'auto'}")
    v = ask("Which one to restore (0 = cancel)", '0')
    if not v.isdigit() or int(v) < 1 or int(v) > len(b): print("   cancelled."); return
    pick = b[int(v)-1]
    print(f"\n   Restoring: {os.path.basename(pick)}")
    print("   Your CURRENT workbook will be backed up first, so this is reversible.")
    if not confirm("   Are you sure?"): print("   cancelled."); return
    loka.backup('before_restore')
    shutil.copy(pick, P)
    print("   restored.")
    loka.refresh_all(do_backup=False)

CMDS = dict(recount=recount, repay=repay, tips=tips, settle=settle, report=report, restore=restore)
if __name__ == '__main__':
    a = sys.argv[1] if len(sys.argv) > 1 else ''
    if a in CMDS:
        try: CMDS[a]()
        except KeyboardInterrupt: print("\n   cancelled.")
    else:
        print(__doc__)
