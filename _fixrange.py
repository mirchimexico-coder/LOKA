# -*- coding: utf-8 -*-
import openpyxl, io, re
from datetime import datetime, date
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
wb=openpyxl.load_workbook(P); dl=wb.worksheets[2]; cap=wb.worksheets[0]
NEW='5000'
changed=[]
# 1) Daily Log H column (expenses SUMIFS) rows 8..max: replace 500 -> 5000 in the ranges
for r in range(8, dl.max_row+1):
    f=dl.cell(r,8).value
    if isinstance(f,str) and 'SUMIFS' in f and '$500' in f:
        nf=f.replace('$F$8:$F$500','$F$8:$F$'+NEW).replace('$B$8:$B$500','$B$8:$B$'+NEW)
        dl.cell(r,8,nf); changed.append(f"DL H{r}")
# 2) Capital Section K references to Expenses/Daily Log with :500
for r in (143,144,145):
    f=cap.cell(r,3).value
    if isinstance(f,str) and '500' in f:
        nf=f.replace('D8:D500','D8:D'+NEW).replace('E8:E500','E8:E'+NEW).replace('G8:G500','G8:G'+NEW).replace('F8:F500','F8:F'+NEW).replace('R8:R500','R8:R'+NEW)
        cap.cell(r,3,nf); changed.append(f"CAP C{r}")
# 3) also scan ALL sheets for any other '!$F$8:$F$500' style Expenses refs
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v=cell.value
            if isinstance(v,str) and v.startswith('=') and 'Expenses' in v and ('F$500' in v or 'B$500' in v or 'F500' in v):
                if ws.title==dl.title: continue
                nf=v.replace('$500',f'${NEW}').replace('F500','F'+NEW).replace('B500','B'+NEW)
                if nf!=v: cell.value=nf; changed.append(f"{ws.title} {cell.coordinate}")
wb.calculation.calcMode='auto'; wb.calculation.fullCalcOnLoad=True
wb.save(P)
io.open(r'C:\LOKA\_out.txt','w',encoding='utf-8',newline='\n').write("changed "+str(len(changed))+" formulas:\n"+"\n".join(changed))
print("changed",len(changed),"formulas; sample H68 now:")
print(dl.cell(68,8).value)
