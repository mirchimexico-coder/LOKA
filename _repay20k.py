# -*- coding: utf-8 -*-
import openpyxl
from copy import copy
from datetime import date
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
wb=openpyxl.load_workbook(P); cap=wb.worksheets[0]
# --- 1) fix C144 typo F50000 -> F5000 ---
if isinstance(cap.cell(144,3).value,str) and 'F50000' in cap.cell(144,3).value:
    cap.cell(144,3, cap.cell(144,3).value.replace('F50000','F5000'))
    print("fixed C144 ->", cap.cell(144,3).value)
# --- 2) reduce ADVANCE OWED (C147) by 20,000 and document the repayment ---
old=float(cap.cell(147,3).value)
assert abs(old-39314.81)<0.01, f"C147 unexpected: {old}"
new=round(old-20000,2)
cap.cell(147,3,new)
cap.cell(147,5,f"Operating costs funded from capital: ledger settlement 30 Jun ($19,314.81) + rent 05 Jul ($20,000) = $39,314.81; LESS $20,000 repaid from operating cash on 21-Jul. Net advance still owed = ${new:,.2f}.")
# add an explicit repayment note row at 149 (if empty)
if cap.cell(149,2).value in (None,''):
    cap.cell(149,2,'  \u21b3 Repaid to Capital from operating cash'); cap.cell(149,2).font=copy(cap.cell(147,2).font)
    c=cap.cell(149,3,-20000.00); c.number_format='$#,##0.00'; c.font=copy(cap.cell(147,3).font)
    cap.cell(149,5,'21-Jul-2026: $20,000 paid back to Capital from restaurant operating cash. Reduces advance owed (C147) and Cash-on-Hand; NOT a P&L expense.')
wb.calculation.calcMode='auto'; wb.calculation.fullCalcOnLoad=True
wb.save(P)
print("C147 advance owed:", old, "->", new)
