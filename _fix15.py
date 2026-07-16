# -*- coding: utf-8 -*-
import openpyxl
from datetime import date
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
wb=openpyxl.load_workbook(P); ex=wb.worksheets[3]; ol=wb['💰 Owner Ledger']; dl=wb.worksheets[2]

# --- 1a) Fix Sam's expense row (477): $199.50 paid=Lohith -> $195 paid=Restaurant (cash) ---
assert "Italpast" in str(ex.cell(477,3).value) and float(ex.cell(477,6).value)==199.50, f"r477 unexpected: {ex.cell(477,3).value} {ex.cell(477,6).value}"
ex.cell(477,6,195.00)
ex.cell(477,7,'Restaurant')
ex.cell(477,8,'Cash')
ex.cell(477,10,"Total $195 (subtotal $199.50 - desc $4.50). Pagado en EFECTIVO por el restaurante. Socio Shashi en el ticket.")

# --- 1b) Remove the Sam's $199.50 entry from Owner Ledger (row 39) ---
assert float(ol.cell(39,4).value or 0)==199.50, f"r39 not 199.50: {ol.cell(39,4).value}"
# shift row 40 (transfer $130) up to 39, then clear 40, and rebuild TOTALS at 40
from copy import copy
def cs(src,dst):
    for c in range(1,9):
        s=ol.cell(src,c); dd=ol.cell(dst,c)
        dd.font=copy(s.font); dd.fill=copy(s.fill); dd.border=copy(s.border); dd.alignment=copy(s.alignment); dd.number_format=s.number_format
# copy the $130 transfer (currently row40) into row39
cs(40,39)
for c in range(1,9): ol.cell(39,c).value=ol.cell(40,c).value
ol.cell(39,6,'=IFERROR(F38+D39-E39,0)')   # fix relative formula for its new row
# rebuild TOTALS at row 40 (was at 41)
cs(41,40)
ol.cell(40,1,'TOTALS & CURRENT BALANCE'); ol.cell(40,4,'=SUM(D4:D39)'); ol.cell(40,5,'=SUM(E4:E39)'); ol.cell(40,6,'=F39')
ol.cell(40,7,'=IF(F39>0,"Restaurant owes Lohith $"&TEXT(F39,"#,##0"),IF(F39<0,"Lohith holds $"&TEXT(ABS(F39),"#,##0"),"Zero balance"))')
for c in range(1,9): ol.cell(41,c).value=None

# --- 2) Add $605 to Jul15 cash: 1500 -> 2105 ---
assert float(dl.cell(64,5).value)==1500, f"row64 cash unexpected: {dl.cell(64,5).value}"
dl.cell(64,5,2105)
prev=dl.cell(64,12).value or ''
dl.cell(64,12, str(prev)+' | +$605 cash added (correction)')

wb.calculation.fullCalcOnLoad=True
wb.save(P)
spent=sum(float(ol.cell(rr,4).value or 0) for rr in range(4,40) if isinstance(ol.cell(rr,4).value,(int,float)))
recv=sum(float(ol.cell(rr,5).value or 0) for rr in range(4,40) if isinstance(ol.cell(rr,5).value,(int,float)))
print("Sam's row477 -> $", ex.cell(477,6).value, "paid=", ex.cell(477,7).value, ex.cell(477,8).value)
print("Jul15 cash ->", dl.cell(64,5).value)
print("Owner Ledger balance now:", round(spent-recv,2))
print("OL r39:", ol.cell(39,2).value, "| r40:", ol.cell(40,1).value)
