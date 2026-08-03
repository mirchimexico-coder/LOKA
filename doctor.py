# -*- coding: utf-8 -*-
"""
LOKA — health check.  Run any time:  py doctor.py
Checks everything that has silently broken before.  All output plain ASCII.
"""
import sys, os, re, zipfile, collections
sys.path.insert(0, r'C:\LOKA')
import openpyxl
from datetime import datetime, date
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
DASH=r'C:\LOKA\dashboard.html'
ok=[]; warn=[]; bad=[]
def _pd(v): return v.date() if isinstance(v,datetime) else (v if isinstance(v,date) else None)

wb=openpyxl.load_workbook(P)
dl=wb.worksheets[2]; ex=wb.worksheets[3]; cap=wb.worksheets[0]
last=ex.max_row
while last>7 and ex.cell(last,3).value is None: last-=1

# 1 capped SUMIFS ranges
# NOTE: a TOTALS row that sums everything above it (e.g. Owner Ledger D56 = SUM(D4:D55))
# is CORRECT, not a cap - it must not be flagged. Only ranges that stop SHORT of the
# data actually matter, so compare the range end against that sheet's last used row.
rng=re.compile(r'\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)')
def _lastused(ws):
    last=ws.max_row
    while last>1 and all(ws.cell(last,c).value in (None,'') for c in range(1,9)): last-=1
    return last
n=0; capped=[]
for ws in wb.worksheets:
    lu=_lastused(ws)
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and v.startswith('='):
                for m in rng.finditer(v):
                    c1,r1,c2,r2=m.group(1),int(m.group(2)),m.group(3),int(m.group(4))
                    if c1!=c2 or (r2-r1)<50: continue
                    # which sheet does this range point at?
                    tgt=ws
                    pre=v[:m.start()]
                    sm=re.findall(r"'([^']+)'!", pre)
                    if sm:
                        try: tgt=wb[sm[-1]]
                        except Exception: tgt=ws
                    tl=_lastused(tgt)
                    # A totals row that sums everything ABOVE itself is correct:
                    #   Owner Ledger D56 = SUM(D4:D55)  -> ends exactly at its own row-1.
                    # Only flag ranges on the SAME sheet that stop short of real data,
                    # or cross-sheet ranges that would miss existing/imminent rows.
                    own_row = c.row
                    if tgt is ws and r2 == own_row - 1:
                        continue                      # classic totals row - fine
                    if r2 <= tl + 20:
                        n+=1; capped.append(f"{ws.title} {c.coordinate} -> {m.group(0)} (data to row {tl})")
(ok if n==0 else bad).append(f"formula ranges that stop short of the data: {n}"
                             + ("" if n==0 else "  -> "+capped[0]+"  FIX: extend it"))

# 2 headroom
(ok if last<4500 else bad).append(f"Expenses last row {last} (cap 5000)")

# 3 text labels starting with '='
lab=0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and v.startswith('=') and ' ' in v[1:] and '(' not in v and '!' not in v \
               and not re.match(r'^[A-Z]+\$?\d', v[1:]) and '+' not in v and '-' not in v:
                lab+=1
(ok if lab==0 else bad).append(f"text labels starting with '=' (#NAME risk): {lab}")

# 4 empty formula caches in Daily Log key columns
x=zipfile.ZipFile(P).read('xl/worksheets/sheet3.xml').decode('utf-8')
empt=[m.group(1) for m in re.finditer(r'<c r="([A-Z]+\d+)"[^>]*>(.*?)</c>', x)
      if '<f' in m.group(2) and re.search(r'<v\s*/>', m.group(2)) and m.group(1)[0] in 'FHIJR']
(ok if not empt else bad).append(f"blank F/H/I/J/R formula caches: {len(empt)}" + ("" if not empt else "  <-- FIX: py loka.py refresh-all"))

# 5 duplicate daily dates / gaps
seen={}; dup=0
for r in range(8,dl.max_row+1):
    d=_pd(dl.cell(r,2).value)
    if isinstance(d,date):
        if d in seen: dup+=1
        seen[d]=r
(ok if dup==0 else bad).append(f"duplicate Daily Log dates: {dup}")

# 6 expenses data quality
q=[]
for r in range(8,last+1):
    if ex.cell(r,3).value in (None,''): continue
    if not isinstance(_pd(ex.cell(r,2).value),date): q.append(f"row {r}: bad date")
    a=ex.cell(r,6).value
    if not isinstance(a,(int,float)): q.append(f"row {r}: non-numeric amount")
    elif a<0: q.append(f"row {r}: negative amount")
    if ex.cell(r,7).value in (None,''): q.append(f"row {r}: blank Paid By")
(ok if not q else bad).append(f"expense data problems: {len(q)}"+("" if not q else "  e.g. "+q[0]))

# 7 category spelling variants (same name differing only by spaces/case)
cats=collections.Counter(str(ex.cell(r,5).value or '').strip() for r in range(8,last+1) if ex.cell(r,5).value)
norm=collections.defaultdict(list)
for c in cats:
    norm[re.sub(r'[\s/]+','',c).lower()].append(c)
variants={k:v for k,v in norm.items() if len(v)>1}
(ok if not variants else warn).append(f"category spelling variants: {len(variants)}"+("" if not variants else "  "+str(list(variants.values()))))

# 8 capital reconciles to 600k
def s(lo,hi): return sum(float(cap.cell(r,4).value or 0) for r in range(lo,hi+1) if isinstance(cap.cell(r,4).value,(int,float)))
try:
    c147=float(cap.cell(147,3).value or 0)
    dep=s(14,30)+(s(56,66)-float(cap.cell(56,4).value or 0))+s(43,51)+s(72,81)+s(114,128)+c147
    withme=545000-(s(56,66)-5000)-s(14,30)-c147
    avail=withme+(180000-145000-s(43,51))+(120000-100000-s(72,81))
    tot=dep+avail
    (ok if abs(tot-600000)<1 else bad).append(f"Capital reconciles to 600,000: {tot:,.2f}")
except Exception as e:
    warn.append("capital check failed: "+str(e))

# 9 dashboard freshness + backup counts
import glob
bdir=r'C:\LOKA\Backup'
xl=len(glob.glob(os.path.join(bdir,'*.xlsx'))); ht=len(glob.glob(os.path.join(bdir,'dashboard_*.html')))
(ok if xl<=31 and ht<=16 else warn).append(f"backups: {xl} xlsx (cap 30), {ht} dashboard html (cap 15)")
stray=[f for f in os.listdir(r'C:\LOKA') if f.startswith('_') and f.endswith('.py')]
(ok if not stray else warn).append(f"stray temp scripts in root: {len(stray)} {stray if stray else ''}")

# 10 propinas table headroom (rows 63-92; fills up and then silently has nowhere to go)
try:
    sp=wb['👥 Staff & Payroll']
    used=sum(1 for r in range(63,93) if sp.cell(r,1).value not in (None,''))
    free=30-used
    msg=f"propinas table: {used}/30 used, {free} slot(s) left"
    if free<=0: bad.append(msg+"  <-- FULL: ask Claude to extend rows 63-92 + the SUM range")
    elif free<=4: warn.append(msg+"  <-- getting full, ask Claude to extend it soon")
    else: ok.append(msg)
except Exception as e:
    warn.append("propinas check failed: "+str(e))

# 11 last recorded day
lastd=max(seen) if seen else None
warn.append(f"last Daily Log date: {lastd}  (row {seen.get(lastd)})") if lastd else None

def _safe(s):
    """Windows console is cp1252 - emoji sheet names crash print(). Strip them."""
    return ''.join(ch for ch in str(s) if ord(ch) < 0x2500).strip()

print("="*62); print("LOKA HEALTH CHECK"); print("="*62)
for m in ok:   print("  [OK]   "+_safe(m))
for m in warn: print("  [note] "+_safe(m))
for m in bad:  print("  [FAIL] "+_safe(m))
print("="*62)
print(("ALL CLEAR" if not bad else f"{len(bad)} PROBLEM(S) - see FAIL lines above"))
