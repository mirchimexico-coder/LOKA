# -*- coding: utf-8 -*-
import openpyxl, zipfile, shutil, re
from datetime import datetime, date
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
wb=openpyxl.load_workbook(P); dl=wb.worksheets[2]; ex=wb.worksheets[3]
def pd(v): return v.date() if isinstance(v,datetime) else (v if isinstance(v,date) else None)
lastx=ex.max_row
while lastx>7 and ex.cell(lastx,3).value is None: lastx-=1
ebd={}
for r in range(8,lastx+1):
    d=pd(ex.cell(r,2).value); a=ex.cell(r,6).value
    if isinstance(d,date) and isinstance(a,(int,float)): ebd[d]=ebd.get(d,0.0)+float(a)
vals={}
for r in range(8, dl.max_row+1):
    d=pd(dl.cell(r,2).value)
    if not isinstance(d,date): continue
    rev=sum(float(dl.cell(r,c).value or 0) for c in (4,5,7,21,25))
    exp=round(ebd.get(d,0.0),2); net=round(rev-exp,2); marg=round(net/rev,6) if rev else 0
    vals[f'F{r}']=rev; vals[f'H{r}']=exp; vals[f'I{r}']=net; vals[f'J{r}']=marg
    vals[f'R{r}']=round(float(dl.cell(r,4).value or 0)*0.0406,2)
src=P; tmp=P+'.tmp'
zin=zipfile.ZipFile(src,'r'); zout=zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED)
inj=0
for item in zin.namelist():
    data=zin.read(item)
    if item=='xl/worksheets/sheet3.xml':
        xml=data.decode('utf-8')
        def repl(m):
            global inj
            ref=m.group('ref'); whole=m.group(0)
            if ref in vals and '<f' in whole:
                newv=f"<v>{vals[ref]}</v>"
                w=re.sub(r'<v\s*/>', newv, whole)
                if w==whole: w=re.sub(r'<v>.*?</v>', newv, whole)
                if '<v' not in w: w=w.replace('</c>', newv+'</c>')
                if w!=whole: inj+=1
                return w
            return whole
        xml=re.sub(r'<c r="(?P<ref>[A-Z]+\d+)"[^>]*>.*?</c>', repl, xml)
        data=xml.encode('utf-8')
    zout.writestr(item,data)
zin.close(); zout.close(); shutil.move(tmp,src)
# verify H67 now
z=zipfile.ZipFile(P); x=z.read('xl/worksheets/sheet3.xml').decode('utf-8')
m=re.search(r'<c r="H67"[^>]*>.*?</c>',x)
print("injected",inj,"cells | H67:",m.group(0) if m else '?')
