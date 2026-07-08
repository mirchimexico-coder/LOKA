# -*- coding: utf-8 -*-
import openpyxl
P=r'C:\LOKA\LOKA_Restaurant_Manager.xlsx'
wb=openpyxl.load_workbook(P); cap=wb.worksheets[0]
# --- Section K: advance owed to capital 19,314.81 -> 39,314.81 ---
old=cap.cell(147,3).value
assert abs(float(old)-19314.81)<0.01, f"unexpected C147 {old}"
cap.cell(147,3, 39314.81)
cap.cell(147,5, "Operating costs funded from capital: ledger settlement 30 Jun ($19,314.81) + rent paid from capital 05 Jul ($20,000). Add future capital-funded operating costs here.")
# --- Section J: link r139 to Section K so they stay in sync, relabel ---
cap.cell(139,2, "\u2212 Operating costs funded from capital (ledger settlement + rent)")
cap.cell(139,4, "=C147")
wb.calculation.fullCalcOnLoad=True
wb.save(P)
print("Section K C147 ->", cap.cell(147,3).value)
print("Section J D139 ->", cap.cell(139,4).value, "| label:", cap.cell(139,2).value)
